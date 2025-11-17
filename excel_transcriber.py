"""
Excel-Based YouTube Audio Transcriber

Структура таблицы:
- Колонка A: Название видео
- Колонка B: Ссылка на YouTube

Использование:
    python excel_transcriber.py           # Автопоиск .xlsx
    python excel_transcriber.py file.xlsx  # Указать файл
"""

import sys
import os

# Fix encoding for Windows console
if sys.platform == 'win32':
    if hasattr(sys.stdout, 'reconfigure'):
        sys.stdout.reconfigure(encoding='utf-8')
        sys.stderr.reconfigure(encoding='utf-8')
    else:
        import codecs
        sys.stdout = codecs.getwriter('utf-8')(sys.stdout.buffer, 'strict')
        sys.stderr = codecs.getwriter('utf-8')(sys.stderr.buffer, 'strict')

import requests
from pytubefix import YouTube
import yt_dlp
import re
import socket
import hashlib
import config
from retry_handler import retry_with_backoff
from openpyxl import load_workbook
from openpyxl.styles import PatternFill


# Цвет для окраски ячеек
GREEN_FILL = PatternFill(start_color="00FF00", end_color="00FF00", fill_type="solid")  # Успех


def clean_filename(filename):
    """Очистка имени файла от недопустимых символов"""
    cleaned = re.sub(r'[<>:"/\\|?*]', '', filename)
    return cleaned.strip('. ')[:100]


def check_dns_availability(hostname: str) -> bool:
    """Проверка доступности DNS"""
    try:
        socket.gethostbyname(hostname)
        return True
    except socket.gaierror:
        return False


def is_row_processed(row) -> bool:
    """
    Проверка, обработана ли строка (окрашена в зеленый).
    
    Args:
        row: Строка из Excel
        
    Returns:
        True если строка уже обработана (зеленая)
    """
    # Проверяем цвет первой ячейки (колонка A)
    cell = row[0]  # Колонка A
    if cell.fill and cell.fill.start_color:
        # Проверяем, является ли цвет зеленым
        color = cell.fill.start_color.rgb if hasattr(cell.fill.start_color, 'rgb') else None
        if color and (color == "0000FF00" or color == "00FF00"):
            return True
    return False


def color_row(worksheet, row_num: int, fill: PatternFill):
    """
    Окрасить строку определенным цветом.
    
    Args:
        worksheet: Лист Excel
        row_num: Номер строки (1-based)
        fill: Цвет заливки
    """
    for col in ['A', 'B']:
        cell = worksheet[f'{col}{row_num}']
        cell.fill = fill


@retry_with_backoff(operation_name="YouTube Download (yt-dlp)")
def download_with_ytdlp_core(youtube_url: str, output_path: str, video_title: str) -> tuple[str, str]:
    """
    Скачивание аудио через yt-dlp с автоматическим retry.
    """
    ydl_info_opts = {
        'quiet': True,
        'no_warnings': True,
    }
    
    with yt_dlp.YoutubeDL(ydl_info_opts) as ydl:
        info = ydl.extract_info(youtube_url, download=False)
        extracted_title = info.get('title', video_title)
        
        if video_title.startswith('video_'):
            video_title = clean_filename(extracted_title)
    
    audio_filename = f"{video_title}.mp4"
    audio_path = os.path.join(output_path, audio_filename)
    
    ydl_opts = {
        'format': 'bestaudio[ext=m4a]/bestaudio/best',
        'outtmpl': audio_path,
        'noplaylist': True,
        'quiet': True,
        'no_warnings': True,
    }
    
    with yt_dlp.YoutubeDL(ydl_opts) as ydl:
        ydl.download([youtube_url])
    
    if not os.path.exists(audio_path):
        raise Exception("Audio file was not created by yt-dlp")
    
    return audio_path, video_title


def download_youtube_audio(youtube_url: str, output_dir: str, suggested_name: str = None) -> tuple[str, str]:
    """
    Скачивание аудио с YouTube (pytubefix → yt-dlp fallback).
    
    Args:
        youtube_url: URL видео
        output_dir: Папка для сохранения
        suggested_name: Предложенное имя из таблицы
        
    Returns:
        Tuple (audio_path, video_title)
    """
    video_title = None
    audio_path = None
    
    # METHOD 1: pytubefix (быстрее)
    try:
        yt = YouTube(youtube_url)
        video_title = clean_filename(suggested_name if suggested_name else yt.title)
        
        audio_stream = yt.streams.filter(only_audio=True).first()
        
        if not audio_stream:
            raise Exception("No audio stream available")
        
        audio_stream.download(output_path=output_dir, filename=f"{video_title}.mp4")
        audio_path = os.path.join(output_dir, f"{video_title}.mp4")
        
        if not os.path.exists(audio_path):
            raise Exception("Audio file not created")
        
        return audio_path, video_title
        
    except Exception as e:
        # METHOD 2: yt-dlp fallback
        try:
            if not video_title:
                video_title = clean_filename(suggested_name) if suggested_name else f"video_{hashlib.md5(youtube_url.encode()).hexdigest()[:8]}"
            
            audio_path, video_title = download_with_ytdlp_core(
                youtube_url, 
                output_dir, 
                video_title
            )
            
            return audio_path, video_title
            
        except Exception as e2:
            print(f"[ERROR] Не удалось скачать аудио: {e2}")
            sys.stdout.flush()
            raise


@retry_with_backoff(operation_name="Audio Transcription")
def transcribe_audio_core(audio_path: str, api_key: str) -> dict:
    """
    Транскрибация аудио через Lemonfox API (с автоматическим retry).
    """
    hostname = "api.lemonfox.ai"
    
    with open(audio_path, 'rb') as audio_file:
        response = requests.post(
            f"https://{hostname}/v1/audio/transcriptions",
            headers={"Authorization": f"Bearer {api_key}"},
            files={"file": audio_file},
            data={"language": "en", "response_format": "json"},
            timeout=(30, 300)
        )
    
    if response.status_code == 200:
        return response.json()
    else:
        response.raise_for_status()
        raise Exception(f"API error: {response.status_code} - {response.text}")


def transcribe_audio_with_retry(audio_path: str, api_key: str) -> dict:
    """Транскрибация с проверкой DNS и автоматическим retry."""
    hostname = "api.lemonfox.ai"
    
    if not check_dns_availability(hostname):
        print("[WARNING] DNS resolution failed - will retry automatically")
        sys.stdout.flush()
    
    return transcribe_audio_core(audio_path, api_key)


def process_video_row(worksheet, row_num: int, video_name: str, video_url: str, workbook_path: str) -> bool:
    """
    Обработка одной строки таблицы: скачивание + транскрибация.
    
    Args:
        worksheet: Лист Excel
        row_num: Номер строки
        video_name: Название видео из колонки A
        video_url: URL из колонки B
        workbook_path: Путь к Excel файлу
        
    Returns:
        True если успешно, False если ошибка
    """
    print(f"\n{'='*60}")
    print(f"[Строка {row_num}] Обработка видео...")
    print(f"Название: {video_name}")
    print(f"URL: {video_url}")
    sys.stdout.flush()
    
    try:
        # Скачиваем аудио
        print("[1/3] Скачивание аудио...")
        sys.stdout.flush()
        
        audio_path, final_name = download_youtube_audio(video_url, config.TEMP_DIR, video_name)
        audio_size_mb = os.path.getsize(audio_path) / (1024 * 1024)
        print(f"[OK] Аудио скачано: {audio_size_mb:.1f} MB")
        sys.stdout.flush()
        
        # Транскрибируем
        print("[2/3] Транскрибация через Lemonfox API...")
        sys.stdout.flush()
        
        result = transcribe_audio_with_retry(audio_path, config.LEMONFOX_API_KEY)
        
        # Сохраняем транскрипцию
        print("[3/3] Сохранение транскрипции...")
        sys.stdout.flush()
        
        output_path = os.path.join(config.OUTPUT_DIR, f"{final_name}.txt")
        with open(output_path, 'w', encoding='utf-8') as f:
            f.write(result['text'])
        
        print(f"[OK] Транскрипция сохранена: {output_path}")
        sys.stdout.flush()
        
        # Удаляем временный аудио файл
        try:
            os.remove(audio_path)
        except:
            pass
        
        # Окрашиваем в зеленый ТОЛЬКО при успехе
        color_row(worksheet, row_num, GREEN_FILL)
        worksheet.parent.save(workbook_path)
        
        return True
        
    except Exception as e:
        print(f"[ERROR] Ошибка обработки видео: {e}")
        sys.stdout.flush()
        
        # НЕ окрашиваем при ошибке - оставляем белым
        return False


def find_excel_file() -> str:
    """
    Поиск Excel файла в текущей директории.
    
    Returns:
        Путь к найденному Excel файлу
    """
    # Ищем все .xlsx файлы в текущей директории
    xlsx_files = [f for f in os.listdir('.') if f.endswith('.xlsx') and not f.startswith('~$')]
    
    if not xlsx_files:
        print("[ERROR] Не найдено ни одного Excel файла (.xlsx) в текущей директории")
        sys.exit(1)
    
    if len(xlsx_files) == 1:
        return xlsx_files[0]
    
    # Если несколько файлов - выводим список для выбора
    print("Найдено несколько Excel файлов:")
    for idx, file in enumerate(xlsx_files, 1):
        print(f"  {idx}. {file}")
    
    try:
        choice = int(input("\nВыберите номер файла: "))
        if 1 <= choice <= len(xlsx_files):
            return xlsx_files[choice - 1]
        else:
            print("[ERROR] Неверный номер")
            sys.exit(1)
    except (ValueError, KeyboardInterrupt):
        print("\n[ERROR] Отменено")
        sys.exit(1)


def main():
    """Главная функция"""
    
    # Если файл указан в аргументах - используем его
    if len(sys.argv) >= 2:
        excel_file = sys.argv[1]
        if not os.path.exists(excel_file):
            print(f"[ERROR] Файл не найден: {excel_file}")
            sys.exit(1)
    else:
        # Иначе ищем .xlsx в текущей директории
        excel_file = find_excel_file()
    
    print("="*60)
    print("Excel-Based YouTube Audio Transcriber")
    print("="*60)
    print(f"Файл: {excel_file}\n")
    sys.stdout.flush()
    
    try:
        # Создаем необходимые папки
        os.makedirs(config.TEMP_DIR, exist_ok=True)
        os.makedirs(config.OUTPUT_DIR, exist_ok=True)
        
        # Загружаем Excel файл
        print("[STEP 1] Загрузка Excel таблицы...")
        sys.stdout.flush()
        
        workbook = load_workbook(excel_file)
        worksheet = workbook.active
        
        # Подсчитываем строки (пропускаем заголовок)
        total_rows = worksheet.max_row
        data_rows = total_rows - 1  # Минус заголовок
        
        print(f"[OK] Найдено строк: {data_rows}")
        print(f"\n[STEP 2] Начинаю обработку видео...")
        sys.stdout.flush()
        
        # Статистика
        successful = 0
        failed = 0
        skipped = 0
        
        # Обрабатываем каждую строку (начиная со 2-й, первая - заголовок)
        for row_num in range(2, total_rows + 1):
            row = list(worksheet[row_num])
            
            # Проверяем, не обработана ли уже строка
            if is_row_processed(row):
                video_name = row[0].value
                print(f"\n[Строка {row_num}] Пропускаю (уже обработана): {video_name}")
                sys.stdout.flush()
                skipped += 1
                continue
            
            # Получаем данные
            video_name = row[0].value  # Колонка A
            video_url = row[1].value   # Колонка B
            
            if not video_name or not video_url:
                print(f"\n[Строка {row_num}] Пропускаю (пустые данные)")
                sys.stdout.flush()
                skipped += 1
                continue
            
            # Обрабатываем видео
            if process_video_row(worksheet, row_num, video_name, video_url, excel_file):
                successful += 1
            else:
                failed += 1
        
        # Финальный отчет
        print(f"\n{'='*60}")
        print(f"ОБРАБОТКА ЗАВЕРШЕНА")
        print(f"{'='*60}")
        print(f"Успешно обработано: {successful}")
        print(f"Ошибок: {failed}")
        print(f"Пропущено (уже обработаны): {skipped}")
        print(f"Транскрипции сохранены в: {config.OUTPUT_DIR}/")
        print(f"{'='*60}")
        sys.stdout.flush()
        
    except Exception as e:
        print(f"\n[FATAL ERROR] {e}")
        import traceback
        traceback.print_exc()
        sys.exit(1)


if __name__ == "__main__":
    main()

